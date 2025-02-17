from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Alignment, Border, Side
from icecream import ic

other_option = {
    "start_row": None,
    "start_column": None,
    "end_row": None,
    "end_column": None
}


def header_builder(worksheet: Worksheet, row_num: int, content: str, h_aligment: str = "center", v_aligment: str = "top") -> Cell:
    """
    Build a header cell in the given worksheet at the given row_num

    :param Worksheet worksheet: The worksheet to build the header cell in
    :param int row_num: The row number to build the header cell at
    :param str content: The content of the header cell
    :param str alignment: The alignment of the header cell, defaults to "center"
    :return Cell: The built header cell
    """
    cell = worksheet.cell(row=row_num, column=1, value=content)
    cell.alignment = Alignment(horizontal=h_aligment, vertical=v_aligment)
    cell.font = Font(bold=True)

    if content == "<<separator>>":
        cell.border = Border(top=Side(style='double'))
        cell.value = ""
    elif content == "<<spacing>>":
        cell.value = ""

    return cell


def cell_builder(worksheet: Worksheet, row_num: int, column_num: int, content: str, h_aligment: str = None, v_aligment: str = None, **other_option) -> Cell:
    """
    Build a cell in the given worksheet at the given row_num and column_num

    :param Worksheet worksheet: The worksheet to build the cell in
    :param int row_num: The row number to build the cell at
    :param int column_num: The column number to build the cell at
    :param str content: The content of the cell
    :param str alignment: The alignment of the cell, defaults to "center"
    :return Cell: The built cell
    """
    cell = worksheet.cell(row=row_num, column=column_num)
    cell.value = content
    cell.alignment = Alignment(horizontal=h_aligment, vertical=v_aligment)
    if other_option:
        if "merge_option" in other_option:
            other_option = other_option["merge_option"]
            ic(other_option)
            worksheet.merge_cells(start_row=other_option["start_row"], start_column=other_option["start_column"],
                                  end_row=other_option["end_row"], end_column=other_option["end_column"])
        elif "border_option" in other_option:
            other_option = other_option["border_option"]
            style_top = Side(other_option["top"]
                             ) if "top" in other_option else None
            style_bottom = Side(
                other_option["bottom"]) if "bottom" in other_option else None
            style_left = Side(
                other_option["left"]) if "left" in other_option else None
            style_right = Side(
                other_option["right"]) if "right" in other_option else None
            cell.border = Border(top=style_top, bottom=style_bottom,
                                 left=style_left, right=style_right)
    return cell


def cell_builder_new(worksheet: Worksheet, row_num: int, column_num: int, content: str, border: str = None) -> Cell:
    """Build a cell in the given worksheet at the given row_num and column_num.

    :param Worksheet worksheet: The worksheet to build the cell in
    :param int row_num: The row number to build the cell at
    :param int column_num: The column number to build the cell at
    :param str content: The content of the cell
    :param str border: The border of the cell, defaults to None
    :return Cell: The built cell
    """
    cell = worksheet.cell(row=row_num, column=column_num)
    cell.value = content
    if border:
        border_chars = list(border.lower())
        border_list = []
        if "t" in border_chars:
            border_list.append(("top", Side(style="thin")))
        if "b" in border_chars:
            border_list.append(("bottom", Side(style="thin")))
        if "l" in border_chars:
            border_list.append(("left", Side(style="thin")))
        if "r" in border_chars:
            border_list.append(("right", Side(style="thin")))
        cell.border = Border(**dict(border_list))

    return cell

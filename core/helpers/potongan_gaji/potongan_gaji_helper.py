import itertools
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from core.excel_helper import cell_builder_new
from core.helper import get_nama_bulan


def generate_potongan(wb: Workbook, title: str, short_name: str, tahun: int, bulan: int, daftar_gaji_pegawai: pd.DataFrame):
    worksheet = wb.copy_worksheet(wb.active)
    worksheet.title = short_name
    bulan_cell = worksheet.cell(row=7, column=1)
    bulan_cell.value = f"Bulan: {get_nama_bulan(bulan)} {tahun}"
    worksheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)
    title_cell = worksheet.cell(row=8, column=1)
    title_cell.value = title
    worksheet.merge_cells(start_row=8, start_column=1, end_row=8, end_column=8)
    row_counter = itertools.count(start=11)

    for index, row in daftar_gaji_pegawai.iterrows():
        generate_empty_row(worksheet, next(row_counter), 7, "TRL")
        generate_potongan_row(worksheet, next(row_counter),
                              row, index+1)
        generate_empty_row(worksheet, next(row_counter), 7, "RLB")


def generate_potongan_row(worksheet: Worksheet, row_num: int, gaji_pegawai_df: pd.DataFrame, urut: int):
    col_num = itertools.count(start=1)

    def build_cell(content: str, is_number: bool = False, border: str = None) -> None:
        cell = cell_builder_new(worksheet, row_num, next(
            col_num), content, border if border else "LR")
        if is_number:
            cell.number_format = "#,##0"

    build_cell(urut)
    build_cell(f"{gaji_pegawai_df['nama']}")
    build_cell(f"{gaji_pegawai_df['nipam']}")
    build_cell(gaji_pegawai_df['penghasilan_bersih'], is_number=True)
    build_cell(0, is_number=True)
    build_cell(0, is_number=True)
    build_cell(gaji_pegawai_df['penghasilan_bersih'], is_number=True)


def generate_empty_row(worksheet: Worksheet, row_num: int, max_col: int, border: str = None) -> None:
    for index in range(max_col):
        cell_builder_new(worksheet, row_num, index+1,
                         "", border if border else "LR")

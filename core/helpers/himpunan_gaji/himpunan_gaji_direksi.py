from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
import itertools
import pandas as pd
from core.databases.gaji_batch_master_proses import get_nilai_komponen, get_total_nilai_komponen
from core.excel_helper import cell_builder
from core.helper import get_nama_bulan


def generate_direksi_sheet(
    workbook: Workbook, year: int, month: int,
    direksi_gaji_df: pd.DataFrame, komponen_gaji_df: pd.DataFrame,
    dirum: pd.DataFrame
) -> None:
    workbook.active = workbook["pegawai"]
    active_sheet = workbook.active
    """Generate a sheet for Direksi."""
    worksheet = workbook.copy_worksheet(active_sheet)
    worksheet.title = "DIREKSI"
    worksheet["A7"] = f"Bulan: {get_nama_bulan(month)} {year}"
    worksheet["A8"] = "DIREKSI"

    row_num = itertools.count(start=12)

    for index, direksi_gaji in direksi_gaji_df.iterrows():
        row_num = itertools.count(
            generate_direksi_row(
                worksheet, next(row_num), index +
                1, direksi_gaji, komponen_gaji_df
            )
        )
    row_num = itertools.count(generate_direksi_footer(
        worksheet, next(row_num), direksi_gaji_df, komponen_gaji_df))

    generate_ttd(worksheet, next(row_num), dirum, year, month)


def generate_direksi_row(
        worksheet: Worksheet,
        start_row: int,
        order_number: int,
        employee: dict,
        salary_components: pd.DataFrame):
    row_counter = itertools.count(start=start_row)
    column_counter = itertools.count(start=1)

    def build_cell(content, is_numeric=False, horizontal_align=None, vertical_align=None):
        cell = cell_builder(
            worksheet=worksheet,
            row_num=start_row,
            column_num=next(column_counter),
            content=content,
            h_aligment=horizontal_align,
            v_aligment=vertical_align,
            border_option={"left": "thin", "right": "thin"}
        )
        if is_numeric:
            cell.number_format = "#,##0"

    build_cell(order_number)
    build_cell("{}{}".format("** " if employee["is_different"] else "", employee["nama"]))
    build_cell(employee["nipam"])
    build_cell("-", halign="center")

    components = [
        ["GP", "0", "TUNJ_JABATAN", "TUNJ_AIR", "POT_PENSIUN", "POT_ASKES", "PENGHASILAN_BERSIH_FINAL"],
        ["", "", "", "JML_JIWA", "TUNJ_SI", "0", "TUNJ_BERAS", "TUNJ_PPH21", "POT_ASTEK", "POT_TKK", "", ""],
        ["", "", "", "", "TUNJ_ANAK", "", "TUNJ_KK", "PENGHASILAN_KOTOR", "SEWA_RUDIN", "POT_PPH21", "", ""],
        ["", "", "", "", "JUMLAH", "", "TUNJ_KESEHATAN", "PEMBULATAN", "POT_JP", "POTONGAN", "", ""]
    ]
    
    for idx, component_list in enumerate(components):
        generate_cell_list(
            worksheet,
            next(row_counter),
            5 if idx == 0 else 1,
            salary_components,
            employee,
            component_list,
            urut=order_number,
            is_first=(idx == 0),
            is_last=(idx == len(components) - 1)
        )

    generate_pemda_title(worksheet, next(row_counter), "Gaji yang telah diterima di PEMDA")

    pemda_values_components = [
        ["0", "0", "0", "0", "0", "0", "0", ""],
        ["0", "0", "0", "0", "0", "0", "", ""],
        ["0", "", "0", "0", "0", "0", "", ""],
        ["0", "", "0", "0", "0", "0", "", ""]
    ]

    for idx, component_list in enumerate(pemda_values_components):
        generate_pemda_value(
            worksheet,
            next(row_counter) - 1,
            employee,
            salary_components,
            component_list,
            is_first=(idx == 0),
            is_last=(idx == len(pemda_values_components) - 1)
        )

    row_counter = itertools.count(start=next(row_counter) - 1)
    generate_pemda_title(worksheet, next(row_counter), "Kekurangan yang harus dibayar PDAM")

    for idx, component_list in enumerate(components):
        generate_pemda_value(
            worksheet,
            next(row_counter) - 1,
            employee,
            salary_components,
            component_list,
            is_first=(idx == 0),
            is_last=(idx == len(components) - 1)
        )

    return next(row_counter) - 1


def generate_cell_list(worksheet: Worksheet, row_num: int, start_col: int,
                       salary_df: pd.DataFrame, row_info: dict, component_list: list,
                       order: int = None, is_first_row=False, is_last_row=False):
    column_index = itertools.count(start=start_col)

    def build_cell(content: str, is_numeric: bool = False, center_align: bool = False):
        cell = cell_builder(
            worksheet,
            row_num,
            next(column_index),
            content,
            h_aligment="center" if center_align else None,
            border_option={"left": "thin", "right": "thin", "bottom": "thin" if is_last_row else None}
        )
        if is_numeric:
            cell.number_format = "#,##0"

    for component in component_list:
        if component == "0":
            build_cell(0, True)
        elif component == "":
            build_cell("")
        elif component == "JUMLAH":
            base_salary = get_nilai_komponen(salary_df, row_info["id"], "GP")
            si_allowance = get_nilai_komponen(salary_df, row_info["id"], "TUNJ_SI")
            child_allowance = get_nilai_komponen(salary_df, row_info["id"], "TUNJ_ANAK")
            total = base_salary + si_allowance + child_allowance
            build_cell(total, True)
        elif component == "JML_JIWA":
            build_cell(f"{row_info['jml_tanggungan']}/{row_info['jml_jiwa']}", center_align=True)
        else:
            build_cell(get_nilai_komponen(salary_df, row_info["id"], component), True)

    if is_first_row:
        build_cell(str(order))
    return row_num + 3


def generate_pemda_title(worksheet: Worksheet, row_num: int, value: str = ""):
    column_index = itertools.count(start=1)

    # Build the first cell without content
    cell_builder(worksheet, row_num, next(column_index), "",
                 border_option={"left": "thin", "right": "thin"})

    # Build the merged cell with bold font
    cell_builder(
        worksheet, row_num, next(column_index),
        value,
        "left", "top",
        border_option={"left": "thin", "right": "thin",
                       "top": "thin", "bottom": "thin"}
    ).font = Font(bold=True)
    worksheet.merge_cells(start_row=row_num, start_column=2,
                          end_row=row_num+3, end_column=4)


def generate_pemda_value(
        worksheet: Worksheet,
        row_num: int,
        row_data: dict,
        salary_components: pd.DataFrame,
        komponen_list: list,
        is_first: bool = False,
        is_last: bool = False):
    col_num = itertools.count(start=5)

    if is_last:
        cell_builder(
            worksheet=worksheet, row_num=row_num, column_num=1, content="",
            border_option={"left": "thin", "right": "thin",
                           "bottom": "thin"}
        )

    def build_cell(value: str, is_number: bool = False) -> None:
        cell = cell_builder(
            worksheet=worksheet,
            row_num=row_num,
            column_num=next(col_num),
            content=value,
            border_option={
                "top": "thin" if is_first else None,
                "left": "thin",
                "right": "thin",
                "bottom": "thin" if is_last else None
            }
        )
        if is_number:
            cell.number_format = "#,##0"

    for index, komponen in enumerate(komponen_list):
        if komponen == "0":
            build_cell(0, True)
        elif komponen == "":
            build_cell("")
        elif komponen == "JUMLAH":
            gaji_pokok = get_nilai_komponen(
                salary_components, row_data["id"], "GP")
            tunj_si = get_nilai_komponen(
                salary_components, row_data["id"], "TUNJ_SI")
            tunj_anak = get_nilai_komponen(
                salary_components, row_data["id"], "TUNJ_ANAK")
            jumlah = gaji_pokok + tunj_si + tunj_anak
            build_cell(jumlah, True)
        else:
            build_cell(get_nilai_komponen(salary_components,
                                          row_data["id"], komponen), True)
    return row_num+1


def generate_direksi_footer(worksheet: Worksheet, row_num: int, row_data: dict, salary_components: pd.DataFrame):
    generate_footer_title(worksheet, row_num, row_data)

    row_counter = itertools.count(start=row_num)

    for index, component_list in enumerate([
        ["GP", "0", "TUNJ_JABATAN", "TUNJ_AIR", "POT_PENSIUN",
            "POT_ASKES", "PENGHASILAN_BERSIH_FINAL", ""],
        ["TUNJ_SI", "0", "TUNJ_BERAS", "TUNJ_PPH21", "POT_ASTEK", "POT_TKK", "", ""],
        ["TUNJ_ANAK", "", "TUNJ_KK", "PENGHASILAN_KOTOR",
            "SEWA_RUDIN", "POT_PPH21", "", ""],
        ["JUMLAH", "", "TUNJ_KESEHATAN", "PEMBULATAN", "POT_JP", "POTONGAN", "", ""]
    ]):
        current_row = next(row_counter)
        generate_footer_values(worksheet, current_row,
                               salary_components, component_list, is_end=index == 3)

    next(row_counter)
    return next(row_counter)


def generate_footer_title(worksheet: Worksheet, row_num: int, row_data: dict):
    total_pegawai = row_data["id"].count()
    dir_column = cell_builder(worksheet=worksheet, row_num=row_num, column_num=1, content="DIREKSI", v_aligment="center", border_option={
        "top": "thin",
        "bottom": "thin",
        "left": "thin",
        "right": "thin"
    })
    dir_column.font = Font(bold=True)
    worksheet.merge_cells(start_row=row_num, start_column=1,
                          end_column=2, end_row=row_num+3)
    jml_column = cell_builder(worksheet=worksheet, row_num=row_num, column_num=3, content=f"{total_pegawai} Pegawai", v_aligment="center", border_option={
        "top": "thin",
        "bottom": "thin",
        "left": "thin",
        "right": "thin"
    })
    jml_column.font = Font(bold=True)
    worksheet.merge_cells(start_row=row_num, start_column=3,
                          end_column=4, end_row=row_num+3)
    return row_num+3


def generate_footer_values(worksheet: Worksheet, row_num: int, salary_components: pd.DataFrame, component_list: list, is_end: bool = False):
    """Generate footer values for Direksi template."""
    col_num = itertools.count(start=5)

    def build_cell(value: str, is_number: bool = False) -> None:
        cell = cell_builder(
            worksheet=worksheet,
            row_num=row_num,
            column_num=next(col_num),
            content=value,
            border_option={
                "left": "thin",
                "right": "thin",
                "bottom": "thin" if is_end else None
            }
        )
        if is_number:
            cell.number_format = "#,##0"

    for component in component_list:
        if component == "0":
            build_cell(0, True)
        elif component == "":
            build_cell("")
        elif component == "JUMLAH":
            total_gaji_pokok = get_total_nilai_komponen(
                salary_components, "GP")
            total_tunj_si = get_total_nilai_komponen(
                salary_components, "TUNJ_SI")
            total_tunj_anak = get_total_nilai_komponen(
                salary_components, "TUNJ_ANAK")
            total = total_gaji_pokok + total_tunj_si + total_tunj_anak
            build_cell(total, True)
        else:
            build_cell(get_total_nilai_komponen(
                salary_components, component), True)


def generate_ttd(worksheet: Worksheet, start_row: int, employee_data: dict, year: int, month: int) -> None:
    row_index = itertools.count(start=start_row)

    def add_cell(content, merge=False):
        row = next(row_index)
        cell_builder(
            worksheet=worksheet,
            row_num=row,
            column_num=10,
            content=content,
            h_aligment="center",
        )
        if merge:
            worksheet.merge_cells(
                start_row=row, start_column=10, end_row=row, end_column=11)

    add_cell(f"Purwokerto,     {get_nama_bulan(month)} {year}", merge=True)
    add_cell("DIREKSI PERUMDAM TIRTA SATRIA", merge=True)
    add_cell("KABUPATEN BANYUMAS", merge=True)
    add_cell("Direktur Umum", merge=True)
    for _ in range(3):
        next(row_index)

    add_cell(employee_data["nama"].values[0], merge=True)
    add_cell(f"NIPAM. {employee_data["nipam"].values[0]}", merge=True)

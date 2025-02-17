import itertools
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
import pandas as pd
from core.enums import STATUS_PEGAWAI
from core.excel_helper import cell_builder
from core.helper import get_nama_bulan


def generate_hg_sheet(workbook: Workbook, organisasi_df: pd.DataFrame, year: int, month: int,
                      gaji_pegawai_df: pd.DataFrame, komponen_gaji_df: pd.DataFrame) -> None:
    """
    Generate HG sheet for given parameters.
    """
    workbook.active = workbook["HG1"]
    worksheet = workbook.copy_worksheet(workbook.active)
    worksheet.title = "HG"

    kontrak_df = gaji_pegawai_df[gaji_pegawai_df["status_pegawai"]
                                 == STATUS_PEGAWAI.KONTRAK.value].reset_index(drop=True)

    # Cabang
    organisasi_cabang_df = organisasi_df[organisasi_df["nama"].str.startswith(
        "CABANG")].reset_index(drop=True)
    pegawai_cabang_df = gaji_pegawai_df[
        (gaji_pegawai_df["kode_organisasi"].str.startswith(tuple(organisasi_cabang_df["kode"].unique().tolist()))) &
        (gaji_pegawai_df["status_pegawai"] != STATUS_PEGAWAI.KONTRAK.value)
    ].reset_index(drop=True)
    komponen_cabang_df = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        pegawai_cabang_df["id"])]
    kontrak_cabang_df = kontrak_df[kontrak_df["kode_organisasi"].str.startswith(
        tuple(organisasi_cabang_df["kode"].unique().tolist()))].reset_index(drop=True)

    # Pusat
    organisasi_pusat_df = organisasi_df[~organisasi_df["id"].isin(
        organisasi_cabang_df["id"])]
    pegawai_pusat_df = gaji_pegawai_df[
        ((gaji_pegawai_df["status_pegawai"] != STATUS_PEGAWAI.KONTRAK.value) &
         (gaji_pegawai_df["kode_organisasi"].str.startswith(tuple(organisasi_pusat_df["kode"].unique().tolist()))) |
         gaji_pegawai_df["level_id"].isin([2, 3, 4]))
    ].reset_index(drop=True)
    kontrak_pusat_df = kontrak_df[kontrak_df["kode_organisasi"].str.startswith(
        tuple(organisasi_pusat_df["kode"].unique().tolist()))].reset_index(drop=True)

    # Filter komponen gaji by cabang and pusat

    komponen_kontrak_cabang_df = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        kontrak_cabang_df["id"])]
    komponen_kontrak_cabang_df.loc[:,
                                   "kode_organisasi"] = komponen_kontrak_cabang_df["kode_organisasi"].str[:3]

    komponen_pusat_df = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        pegawai_pusat_df["id"])]
    komponen_kontrak_pusat_df = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        kontrak_pusat_df["id"])]

    # Generate rows in the worksheet
    row_counter = itertools.count(start=7)
    order_counter = itertools.count(start=1)
    next_row, next_order = generate_gapok_row(
        worksheet,
        next(row_counter),
        komponen_pusat_df,
        komponen_cabang_df,
        komponen_kontrak_pusat_df,
        komponen_kontrak_cabang_df,
        next(order_counter)
    )
    row_counter = itertools.count(start=next_row)
    order_counter = itertools.count(start=next_order)

    next_row, next_order = generate_tunjangan_row(
        worksheet,
        next(row_counter),
        komponen_pusat_df,
        komponen_cabang_df,
        next(order_counter),
    )
    row_counter = itertools.count(start=next_row)
    order_counter = itertools.count(start=next_order)

    penghasilan_kotor_row = next(row_counter)
    next_row = generate_penghasilan_kotor_row(worksheet, penghasilan_kotor_row)
    row_counter = itertools.count(start=next_row)

    start_row_potongan = next(row_counter)
    next_row = generate_potongan_row(
        worksheet,
        start_row_potongan,
        organisasi_df,
        gaji_pegawai_df,
        komponen_gaji_df,
        next(order_counter),
    )
    row_counter = itertools.count(start=next_row)

    next_row = generate_jumlah_potongan_row(
        worksheet, start_row_potongan, next(row_counter))
    row_counter = itertools.count(start=next_row)

    next_row = generate_penghasilan_bersih_row(
        worksheet,
        next(row_counter),
        penghasilan_kotor_row
    )
    row_counter = itertools.count(start=next_row)

    generate_ttd(worksheet, next(row_counter), gaji_pegawai_df, year, month)



def generate_gapok_row(
    worksheet: Worksheet,
    start_row: int,
    pusat_gaji: pd.DataFrame,
    cabang_gaji: pd.DataFrame,
    pusat_kontrak_gaji: pd.DataFrame,
    cabang_kontrak_gaji: pd.DataFrame,
    start_urut: int,
) -> tuple[int, int]:
    """
    Generate HG sheet rows for gapok data.
    """
    row_num = itertools.count(start=start_row)
    urut = itertools.count(start=start_urut)
    # row Gaji Pokok
    generate_row(
        worksheet,
        next(row_num),
        pusat_gaji,
        cabang_gaji,
        "GP",
        "Gaji Pokok",
        next(urut),
    )
    # row Kontrak
    generate_row(
        worksheet,
        next(row_num),
        pusat_kontrak_gaji,
        cabang_kontrak_gaji,
        "GP",
        "Honor Tenaga Kontrak",
        next(urut),
    )
    generate_empty_row(worksheet, next(row_num), 9)
    generate_empty_row(worksheet, next(row_num), 9)

    return next(row_num), next(urut)


def generate_tunjangan_row(
    worksheet: Worksheet, start_row: int, central_gaji: pd.DataFrame, branch_gaji: pd.DataFrame, start_order: int
) -> tuple[int, int]:
    """
    Generate HG sheet rows for tunjangan data.
    """
    row_num = itertools.count(start=start_row)
    order = itertools.count(start=start_order)
    empty_row_num = next(row_num)
    generate_empty_row(worksheet, empty_row_num, 9)
    tunjangan_cell = worksheet.cell(row=empty_row_num, column=2)
    tunjangan_cell.value = "Tunjangan:"
    tunjangan_cell.font = Font(bold=True)

    def build_row(kode: str, description: str) -> None:
        generate_row(
            worksheet,
            next(row_num),
            central_gaji,
            branch_gaji,
            kode,
            description,
            next(order),
        )

    # row Tunjangan
    build_row("TUNJ_SI", "Istri/Suami")
    build_row("TUNJ_ANAK", "Anak")
    build_row("TUNJ_JABATAN", "Jabatan / Pelaksana")
    build_row("TUNJ_BERAS", "Beras")
    build_row("TUNJ_KK", "Kegiatan Kerja")
    build_row("TUNJ_AIR", "Air")
    build_row("TUNJ_PPH21", "P.Ph. Pasal 21")
    build_row("PEMBULATAN", "Pembulatan")
    generate_empty_row(worksheet, next(row_num), 9)
    return next(row_num), next(order)


def generate_penghasilan_kotor_row(worksheet: Worksheet, row_num: int) -> int:
    column_index = itertools.count(start=1)

    def build_cell(content: str, is_number: bool = False) -> None:
        """
        Build a cell in the worksheet.
        """
        cell = cell_builder(
            worksheet,
            row_num,
            next(column_index),
            content,
            border_option={"left": "thin", "right": "thin", "bottom": "thin"},
        )
        cell.font = Font(bold=True)
        if is_number:
            cell.number_format = "#,##0"
        return cell

    build_cell("")
    build_cell("Jumlah Penghasilan Kotor")
    build_cell(f"=SUM(C{7}:C{row_num-1})", True)
    build_cell(f"=SUM(D{7}:D{row_num-1})", True)
    build_cell(f"=SUM(E{7}:E{row_num-1})", True)
    build_cell(f"=SUM(F{7}:F{row_num-1})", True)
    build_cell(f"=SUM(G{7}:G{row_num-1})", True)
    build_cell(f"=SUM(H{7}:H{row_num-1})", True)
    build_cell(f"=SUM(I{7}:I{row_num-1})", True)

    return row_num+1


def generate_potongan_row(worksheet: Worksheet, row_num: int, organisasi_df: pd.DataFrame, gaji_pegawai_df: pd.DataFrame, komponen_gaji_df: pd.DataFrame, start_urut: int):
    row_num = itertools.count(start=row_num)
    urut = itertools.count(start=start_urut)
    generate_empty_row(worksheet, next(row_num), 9)
    title_row_num = next(row_num)
    generate_empty_row(worksheet, title_row_num, 9)
    title_cell = worksheet.cell(row=title_row_num, column=2)
    title_cell.value = "Potongan-potongan:"
    title_cell.font = Font(bold=True)

    # Himpunan Gaji Cabang
    org_cabang = organisasi_df[organisasi_df["nama"].str.startswith("CABANG")]
    gaji_pegawai_cabang = gaji_pegawai_df[(gaji_pegawai_df["kode_organisasi"].str.startswith(tuple(org_cabang["kode"].tolist())) &
                                           (gaji_pegawai_df["status_pegawai"] != STATUS_PEGAWAI.KONTRAK.value))]
    komponen_pegawai_cabang = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        gaji_pegawai_cabang["id"].tolist())]
    gaji_kontrak_cabang = gaji_pegawai_df[(gaji_pegawai_df["kode_organisasi"].str.startswith(tuple(org_cabang["kode"].tolist())) &
                                           (gaji_pegawai_df["status_pegawai"] == STATUS_PEGAWAI.KONTRAK.value))]
    komponen_kontrak_cabang = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        gaji_kontrak_cabang["id"].tolist())]

    # Himpunan Gaji Pusat
    org_pusat = organisasi_df[~organisasi_df["nama"].str.startswith("CABANG")]
    gaji_pegawai_pusat = gaji_pegawai_df[
        (
            (gaji_pegawai_df["kode_organisasi"].str.startswith(tuple(org_pusat["kode"].tolist())) &
             (gaji_pegawai_df["status_pegawai"] != STATUS_PEGAWAI.KONTRAK.value))
        )
        | (gaji_pegawai_df["level_id"].isin([2, 3, 4]))
    ]
    komponen_pegawai_pusat = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        gaji_pegawai_pusat["id"].tolist())]
    gaji_kontrak_pusat = gaji_pegawai_df[(gaji_pegawai_df["kode_organisasi"].str.startswith(tuple(org_pusat["kode"].tolist())) &
                                          (gaji_pegawai_df["status_pegawai"] == STATUS_PEGAWAI.KONTRAK.value))]
    komponen_kontrak_pusat = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        gaji_kontrak_pusat["id"].tolist())]

    def build_row(komponen_pusat: pd.DataFrame, komponen_cabang: pd.DataFrame, kode: str, description: str) -> None:
        generate_row(
            worksheet,
            next(row_num),
            komponen_pusat,
            komponen_cabang,
            kode,
            description,
            next(urut),
        )

    build_row(komponen_pegawai_pusat, komponen_pegawai_cabang,
              "POT_PENSIUN", "Iuran Pensiun")
    generate_row(
        worksheet,
        next(row_num),
        komponen_pegawai_pusat,
        komponen_pegawai_cabang,
        "POT_ASTEK",
        "Iuran Jamsostek",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_kontrak_pusat,
        komponen_kontrak_cabang,
        "POT_ASTEK",
        "Iuran Jamsostek T. Kontrak",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_pegawai_pusat,
        komponen_pegawai_cabang,
        "POT_RUDIN",
        "Sewa Rumah",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_pegawai_pusat,
        komponen_pegawai_cabang,
        "POT_JP",
        "Iuran Jamsostek (JPn)",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_kontrak_pusat,
        komponen_kontrak_cabang,
        "POT_JP",
        "Iuran Jamsostek (JPn) T. Kontrak",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_pegawai_pusat,
        komponen_pegawai_cabang,
        "POT_ASKES",
        "Jmn Kes Nasional",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_kontrak_pusat,
        komponen_kontrak_cabang,
        "POT_ASKES",
        "Jmn Kes Nasional T. Kontrak",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_pegawai_pusat,
        komponen_pegawai_cabang,
        "POT_KK",
        "TKK",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_pegawai_pusat,
        komponen_pegawai_cabang,
        "HONOR",
        "Honor",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_pegawai_pusat,
        komponen_pegawai_cabang,
        "POT_PPH21",
        "P.Ph. Pasal 21",
        next(urut),
    )
    generate_row(
        worksheet,
        next(row_num),
        komponen_pegawai_pusat,
        komponen_pegawai_cabang,
        "POT_PPH21_PEMBINA",
        "P.Ph. Pasal 21 Bd Pembinda dll",
        next(urut),
    )
    return next(row_num)


def generate_jumlah_potongan_row(worksheet: Worksheet, start_row: int, current_row: int):
    column_index = itertools.count(start=1)

    def build_cell(content: str, is_number: bool = False) -> None:
        """
        Build a cell in the worksheet.
        """
        cell = cell_builder(
            worksheet,
            current_row,
            next(column_index),
            content,
            border_option={"left": "thin", "right": "thin", "bottom": "thin"},
        )
        cell.font = Font(bold=True)
        if is_number:
            cell.number_format = "#,##0"

    build_cell("")
    build_cell("Jumlah Potongan")
    build_cell(f"=SUM(C{start_row}:C{current_row-1})", True)
    build_cell(f"=SUM(D{start_row}:D{current_row-1})", True)
    build_cell(f"=SUM(E{start_row}:E{current_row-1})", True)
    build_cell(f"=SUM(F{start_row}:F{current_row-1})", True)
    build_cell(f"=SUM(G{start_row}:G{current_row-1})", True)
    build_cell(f"=SUM(H{start_row}:H{current_row-1})", True)
    build_cell(f"=SUM(I{start_row}:I{current_row-1})", True)

    return current_row + 1


def generate_penghasilan_bersih_row(worksheet: Worksheet, row_num: int, penghasilan_kotor_row: int) -> int:
    column_index = itertools.count(start=1)

    def build_cell(content: str, is_number: bool = False) -> None:
        """
        Build a cell in the worksheet.
        """
        cell = cell_builder(
            worksheet,
            row_num,
            next(column_index),
            content,
            border_option={"left": "thin", "right": "thin", "bottom": "thin"},
        )
        cell.font = Font(bold=True)
        if is_number:
            cell.number_format = "#,##0"

    build_cell("")
    build_cell("Jumlah Penghasilan Bersih")
    build_cell(f"=C{penghasilan_kotor_row}-C{row_num-1}", True)
    build_cell(f"=D{penghasilan_kotor_row}-D{row_num-1}", True)
    build_cell(f"=E{penghasilan_kotor_row}-E{row_num-1}", True)
    build_cell(f"=F{penghasilan_kotor_row}-F{row_num-1}", True)
    build_cell(f"=G{penghasilan_kotor_row}-G{row_num-1}", True)
    build_cell(f"=H{penghasilan_kotor_row}-H{row_num-1}", True)
    build_cell(f"=I{penghasilan_kotor_row}-I{row_num-1}", True)

    return row_num + 1

def generate_ttd(worksheet: Worksheet, start_row: int, employee_data: dict, year:int, month:int) -> None:
    """Generate TTD in HG sheet"""
    row_index = itertools.count(start=start_row)
    next(row_index)
    next(row_index)

    # Tanggal
    tanggal_row = next(row_index)
    tanggal_cell = worksheet.cell(row=tanggal_row, column=6)
    tanggal_cell.value = f"Purwokerto,      {get_nama_bulan(month)} {year}"
    tanggal_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=tanggal_row, start_column=6, end_row=tanggal_row, end_column=8
    )

    # Title Direksi
    title_direksi_row = next(row_index)
    title_direksi_cell = worksheet.cell(row=title_direksi_row, column=6)
    title_direksi_cell.value = "DIREKSI PERUMDAM TIRTA SATRIA"
    title_direksi_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=title_direksi_row, start_column=6, end_row=title_direksi_row, end_column=8
    )

    # Title Mengetahui & Kabupaten
    title_mengetahui_row = next(row_index)
    title_mengetahui_cell = worksheet.cell(row=title_mengetahui_row, column=2)
    title_mengetahui_cell.value = "Mengetahui,"
    title_mengetahui_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=title_mengetahui_row, start_column=2, end_row=title_mengetahui_row, end_column=4
    )
    title_kabupaten_cell = worksheet.cell(row=title_mengetahui_row, column=6)
    title_kabupaten_cell.value = "Kabupaten Banyumas"
    title_kabupaten_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=title_mengetahui_row, start_column=6, end_row=title_mengetahui_row, end_column=8
    )

    # Title Direktur
    title_direktur_row = next(row_index)
    title_direktur_cell = worksheet.cell(row=title_direktur_row, column=2)
    title_direktur_cell.value = "Direktur Utama"
    title_direktur_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=title_direktur_row, start_column=2, end_row=title_direktur_row, end_column=4
    )
    title_direktur_cell = worksheet.cell(row=title_direktur_row, column=6)
    title_direktur_cell.value = "Direktur Umum"
    title_direktur_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=title_direktur_row, start_column=6, end_row=title_direktur_row, end_column=8
    )
    next(row_index)
    next(row_index)
    next(row_index)

    direktur_utama = employee_data[employee_data["level_id"] == 2].reset_index(drop=True)
    direktur_umum = employee_data[employee_data["level_id"] == 4].reset_index(drop=True)

    # Nama Direktur
    nama_direktur_row = next(row_index)
    nama_direktur_cell = worksheet.cell(row=nama_direktur_row, column=2)
    nama_direktur_cell.value = direktur_utama["nama"].values[0]
    nama_direktur_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=nama_direktur_row, start_column=2, end_row=nama_direktur_row, end_column=4
    )
    nama_direktur_cell = worksheet.cell(row=nama_direktur_row, column=6)
    nama_direktur_cell.value = direktur_umum["nama"].values[0]
    nama_direktur_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=nama_direktur_row, start_column=6, end_row=nama_direktur_row, end_column=8
    )

    # Nipam Direktur
    nipam_direktur_row = next(row_index)
    nipam_direktur_cell = worksheet.cell(row=nipam_direktur_row, column=2)
    nipam_direktur_cell.value = "NIPAM. " + direktur_utama["nipam"].values[0]
    nipam_direktur_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=nipam_direktur_row, start_column=2, end_row=nipam_direktur_row, end_column=4
    )
    nipam_direktur_cell = worksheet.cell(row=nipam_direktur_row, column=6)
    nipam_direktur_cell.value = "NIPAM. " + direktur_umum["nipam"].values[0]
    nipam_direktur_cell.alignment = Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=nipam_direktur_row, start_column=6, end_row=nipam_direktur_row, end_column=8
    )


def generate_row(
        worksheet: Worksheet,
        row_number: int,
        central_gaji_components: pd.DataFrame,
        branch_gaji_components: pd.DataFrame,
        component_code: str,
        component_description: str,
        order: int = None,
        is_bold: bool = False
) -> None:
    """
    Generate a row in HG sheet for a given component.
    """
    column_index = itertools.count(start=1)

    def build_cell(content: str, is_number: bool = False) -> None:
        """
        Build a cell in the worksheet.
        """
        cell = cell_builder(
            worksheet,
            row_number,
            next(column_index),
            content,
            border_option={"left": "thin", "right": "thin", "bottom": "thin"},
        )
        if is_number:
            cell.number_format = "#,##0"
        if is_bold:
            cell.font = Font(bold=True)
        return cell

    build_cell(order)
    build_cell(component_description)

    central_filtered = filter_gaji_components_by_code(
        central_gaji_components, component_code)
    branch_filtered = filter_gaji_components_by_code(
        branch_gaji_components, component_code)

    # Generate Gaji Pokok Pusat
    build_cell(central_filtered["nilai"].sum(), True)
    # Generate Gaji Pokok Cabang
    build_cell(
        branch_filtered[branch_filtered["kode_organisasi"].str.startswith(
            "1.4")]["nilai"].sum(),  # Cabang Pwt-1
        True,
    )
    build_cell(
        branch_filtered[branch_filtered["kode_organisasi"].str.startswith(
            "1.5")]["nilai"].sum(),  # Cabang Pwt-2
        True,
    )
    build_cell(
        branch_filtered[branch_filtered["kode_organisasi"].str.startswith(
            "1.8")]["nilai"].sum(),  # Cabang Ajb
        True,
    )
    build_cell(
        branch_filtered[branch_filtered["kode_organisasi"].str.startswith(
            "1.7")]["nilai"].sum(),  # Cabang Wgn
        True,
    )
    build_cell(
        branch_filtered[branch_filtered["kode_organisasi"].str.startswith(
            "1.6")]["nilai"].sum(),  # Cabang Bms
        True,
    )
    # Generate Gaji Pokok Total
    build_cell(
        central_filtered["nilai"].sum() + branch_filtered["nilai"].sum(), True)


def generate_empty_row(worksheet: Worksheet, row_num: int, col_count: int = 1) -> None:
    """Generate an empty row in the given worksheet at the given row_num."""
    for col_index in range(1, col_count + 1):
        cell_builder(
            worksheet, row_num, col_index, "",
            border_option={"left": "thin", "right": "thin", "bottom": "thin"}
        )


def filter_gaji_components_by_code(gaji_components_df: pd.DataFrame, code: str) -> pd.DataFrame:
    """Filter gaji components by code."""
    return gaji_components_df[gaji_components_df["kode"] == code].reset_index(drop=True)

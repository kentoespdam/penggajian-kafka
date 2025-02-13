import itertools
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
import pandas as pd

from core.databases.gaji_batch_master_proses import get_total_nilai_komponen
from core.enums import STATUS_PEGAWAI
from core.excel_helper import cell_builder
from core.helper import get_nama_bulan


def generate_hgpkp_sheet(
    workbook: Workbook, organisasi_df: pd.DataFrame, year: int, month: int,
    gaji_pegawai_df: pd.DataFrame, komponen_gaji_df: pd.DataFrame
) -> None:
    workbook.active = workbook["HGPKP1"]
    worksheet = workbook.copy_worksheet(workbook.active)
    worksheet.title = "HGPKP"
    worksheet.cell(row=7, column=2,
                   value=f"Bulan: {get_nama_bulan(month)} {year}")

    row_num = itertools.count(start=16)
    urut = itertools.count(start=2)

    # Generate direksi row
    direksi_ids = gaji_pegawai_df[gaji_pegawai_df["level_id"].isin(
        [2, 3, 4])]["id"].tolist()
    komponen_gaji_direksi = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        direksi_ids)].reset_index(drop=True)
    row_num = itertools.count(generate_row(worksheet, next(
        row_num), komponen_gaji_direksi, len(direksi_ids), "DIREKSI", next(urut)))

    # Generate organisasi row
    organisasi_wt_cabang = organisasi_df[~organisasi_df["nama"].str.startswith(
        "CABANG")].reset_index(drop=True)
    for _, organisasi in organisasi_wt_cabang.iterrows():
        pegawai_ids = gaji_pegawai_df[(gaji_pegawai_df["kode_organisasi"].str.startswith(organisasi["kode"])) &
                                      (gaji_pegawai_df["status_pegawai"] != STATUS_PEGAWAI.KONTRAK.value)]["id"].tolist()

        komponen_gaji_organisasi = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
            pegawai_ids)].reset_index(drop=True)
        row_num = itertools.count(generate_row(worksheet, next(
            row_num), komponen_gaji_organisasi, len(pegawai_ids), organisasi["nama"], next(urut)))

    # Generate footer
    all_pegawai = gaji_pegawai_df[(gaji_pegawai_df["status_pegawai"] != STATUS_PEGAWAI.KONTRAK.value) &
                                  (gaji_pegawai_df["kode_organisasi"].str.startswith(tuple(organisasi_wt_cabang["kode"].unique().tolist())) |
                                   (gaji_pegawai_df["level_id"].isin([2, 3, 4])))]
    all_pegawai_ids = all_pegawai["id"].tolist()
    komponen_gaji_pegawai = komponen_gaji_df[komponen_gaji_df["master_batch_id"].isin(
        all_pegawai_ids)].reset_index(drop=True)
    row_num = itertools.count(generate_row(worksheet, next(
        row_num), komponen_gaji_pegawai, len(all_pegawai_ids), row_name="JUMLAH"))
    current_row_num = next(row_num)
    worksheet.merge_cells(start_row=current_row_num-4,
                          start_column=1, end_row=current_row_num-1, end_column=2)
    jml_cell = worksheet.cell(row=current_row_num-4, column=1)
    jml_cell.alignment = Alignment(horizontal="center", vertical="center")
    jml_cell.font = Font(bold=True)


def generate_row(
        worksheet: Worksheet,
        row_num: int,
        salary_components: pd.DataFrame,
        jml_pegawai: int,
        row_name: str,
        urut: int = None
):
    row_counter = itertools.count(start=row_num)
    column_index = itertools.count(start=1)

    def build_cell(value, is_number=False, halign=None, valign=None):
        cell = cell_builder(
            worksheet=worksheet, row_num=row_num, column_num=next(column_index), content=value,
            h_aligment=halign, v_aligment=valign, border_option={
                "left": "thin", "right": "thin"
            }
        )
        if is_number:
            cell.number_format = "#,##0"
        return cell

    if urut:
        build_cell(urut, True)
        build_cell(row_name)
    else:
        build_cell(row_name)
        next(column_index)
    build_cell(jml_pegawai, halign="center")
    for index, component_list in enumerate([
        ["GP", "0", "TUNJ_JABATAN", "TUNJ_AIR", "POT_PENSIUN",
            "POT_ASKES", "PENGHASILAN_BERSIH_FINAL", ""],
        ["TUNJ_SI", "0", "TUNJ_BERAS",
            "TUNJ_PPH21", "POT_ASTEK", "POT_TKK", "", ""],
        ["TUNJ_ANAK", "", "TUNJ_KK", "PENGHASILAN_KOTOR",
            "SEWA_RUDIN", "POT_PPH21", "", ""],
        ["JUMLAH", "", "TUNJ_KESEHATAN",
            "PEMBULATAN", "POT_JP", "POTONGAN", "", ""]
    ]):
        current_row = next(row_counter)
        generate_cell_list(worksheet, current_row,
                           salary_components, component_list, is_first=index == 0, is_last=index == 3)

    return next(row_counter)


def generate_cell_list(worksheet: Worksheet, row_num: int,
                       salary_components: pd.DataFrame, komponen_list: list,
                       is_first=False, is_last=False):
    col_num = itertools.count(start=4 if is_first else 1)
    border_top = False

    def build_cell(
        value: str, is_number: bool = False
    ):
        cell = cell_builder(
            worksheet=worksheet,
            row_num=row_num,
            column_num=next(col_num),
            content=value,
            border_option={
                "top": "thin" if border_top else None,
                "left": "thin",
                "right": "thin",
                "bottom": "thin" if is_last else None
            }
        )
        if is_number:
            cell.number_format = "#,##0"
        return cell

    if not is_first:
        for _ in range(3):
            build_cell("")

    for component in komponen_list:
        border_top = False
        if component == "POTONGAN":
            border_top = True
        if component == "0":
            build_cell(0, True)
        elif component == "":
            build_cell("")
        elif component == "JUMLAH":
            gaji_pokok = get_total_nilai_komponen(
                salary_components, "GP")
            tunj_si = get_total_nilai_komponen(
                salary_components, "TUNJ_SI")
            tunj_anak = get_total_nilai_komponen(
                salary_components, "TUNJ_ANAK")
            jumlah = gaji_pokok + tunj_si + tunj_anak
            build_cell(jumlah, True)
        else:
            cur_cell = build_cell(get_total_nilai_komponen(
                salary_components, component), True)


def generate_footer(worksheet: Worksheet, row_num: int):
    col_num = itertools.count(start=1)

    def build_cell(value: str, is_number: bool = False, h_align: str = None):
        cell = cell_builder(
            worksheet=worksheet,
            row_num=row_num,
            column_num=next(col_num),
            content=value,
            h_aligment=h_align
        )
        if is_number:
            cell.number_format = "#,##0"

        return cell

    build_cell("Total Sampai Halaman Ini", h_align="center")
    worksheet.merge_cells(start_row=row_num, start_column=1,
                          end_row=row_num+4, end_column=2)

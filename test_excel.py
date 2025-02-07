import itertools
from core.databases.organisasi import fetch_organisasi_by_level
from core.excel_helper import cell_builder, header_builder
from core.helper import get_nama_bulan
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from icecream import ic


def header_value_list(root_batch_id: str):
    periode = root_batch_id.split("-")[0]
    tahun = int(periode[0:4])
    bulan = int(periode[4:6])
    return [
        "PEMERINTAH KABUPATEN BANYUMAS",
        "PERUSAHAAN DAERAH AIR MINUM TIRTA SATRIA",
        "Jalan Prod. Dr. Suharso No. 52 Purwokerto 53114",
        "Telp. & Fax. (0281) - 632324",
        "<<separator>>",
        "Daftar: Gaji Pegawai",
        "Bulan: {} {}".format(get_nama_bulan(bulan), tahun)
    ]


def table_head_builder(ws: Worksheet, row_num: int):
    column_num = itertools.count(1)
    current_col_num = next(column_num)
    cell_builder(worksheet=ws, row_num=row_num, column_num=current_col_num, content="No.", h_aligment="center", v_aligment="center", merge_option={
        "start_row": row_num,
        "start_column": current_col_num,
        "end_row": row_num+1,
        "end_column": current_col_num
    })
    current_col_num = next(column_num)
    cell_builder(worksheet=ws, row_num=row_num, column_num=current_col_num, content="Nama", h_aligment="center", v_aligment="center", merge_option={
        "start_row": row_num,
        "start_column": current_col_num,
        "end_row": row_num+1,
        "end_column": current_col_num
    })
    current_col_num = next(column_num)
    cell_builder(worksheet=ws, row_num=row_num, column_num=current_col_num, content="Nipam", h_aligment="center", v_aligment="center", merge_option={
        "start_row": row_num,
        "start_column": current_col_num,
        "end_row": row_num+1,
        "end_column": current_col_num
    })
    current_col_num = next(column_num)
    cell_builder(worksheet=ws, row_num=row_num, column_num=current_col_num, content="Gol/Ruang \nJml. Jiwa", h_aligment="center", v_aligment="center", merge_option={
        "start_row": row_num,
        "start_column": current_col_num,
        "end_row": row_num+1,
        "end_column": current_col_num
    })
    current_col_num = next(column_num)
    cell_builder(worksheet=ws, row_num=row_num, column_num=current_col_num, content="Penghasilan", h_aligment="center", v_aligment="center", merge_option={
        "start_row": row_num,
        "start_column": current_col_num,
        "end_row": row_num,
        "end_column": current_col_num+3
    })
    next(column_num)
    next(column_num)
    next(column_num)
    current_col_num = next(column_num)
    cell_builder(worksheet=ws, row_num=row_num, column_num=current_col_num, content="Potongan", h_aligment="center", v_aligment="center", merge_option={
        "start_row": row_num,
        "start_column": current_col_num,
        "end_row": row_num,
        "end_column": current_col_num+1
    })
    next(column_num)
    current_col_num = next(column_num)
    cell_builder(worksheet=ws, row_num=row_num, column_num=current_col_num, content="Jumlah \nPenghasilan \nBersih", h_aligment="center", v_aligment="center", merge_option={
        "start_row": row_num,
        "start_column": current_col_num,
        "end_row": row_num+1,
        "end_column": current_col_num
    })
    current_col_num = next(column_num)
    cell_builder(worksheet=ws, row_num=row_num, column_num=current_col_num, content="Tanda \nTangan", h_aligment="center", v_aligment="center", merge_option={
        "start_row": row_num,
        "start_column": current_col_num,
        "end_row": row_num+1,
        "end_column": current_col_num
    })


def main():
    organisasi_list = pd.DataFrame(fetch_organisasi_by_level(4))
    organisasi_list.loc[:, "short_name"] = organisasi_list["nama"].apply(
        lambda x: x.replace("CABANG", "").replace("UNIT BISNIS", "").replace("BAG.", "").strip())

    write_excel(organisasi_list, header_value_list("202402-001"))


def write_excel(organisasi: pd.DataFrame, header: list):
    wb = Workbook()

    row_num = itertools.count(1)
    ws = wb.active
    for index, value in enumerate(header, 1):
        current_row = next(row_num)
        header_builder(ws, current_row, value,
                       h_aligment="center" if index <= 6 else "left")
        if index <= 6:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=6)

    header_builder(ws, next(row_num),
                   organisasi["short_name"].values[0], "left")
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=2)

    table_head_builder(ws, next(row_num))
    wb.save("test.xlsx")


if __name__ == "__main__":
    main()

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.databases.organisasi import fetch_organisasi_by_level
from core.helper import get_nama_bulan
from icecream import ic
import pandas as pd


def main(root_batch_id: str):
    periode = root_batch_id.split("-")[0]
    tahun = int(periode[0:4])
    bulan = int(periode[4:6])
    wb = load_workbook("excel_template/daftar_gaji_template.xlsx")
    # copy sheet_template to wb
    ws = wb.active
    ws.title = "DIREKSI"
    ws["A7"] = "Bulan: {} {}".format(get_nama_bulan(bulan), tahun)
    ws["A8"] = "DIREKSI"

    organisasi_list = pd.DataFrame(fetch_organisasi_by_level([4]))
    organisasi_list.loc[:, "short_name"] = organisasi_list["nama"].apply(
        lambda x: x.replace("CABANG", "").replace("UNIT BISNIS", "").replace("BAG.", "").strip())

    generate_sheet_organisasi(wb, ws, organisasi_list, tahun, bulan)
    wb.save("test_template.xlsx")


def generate_sheet_organisasi(wb: Workbook, ws: Worksheet, organisasi_list: pd.DataFrame, tahun: int, bulan: int):
    for _, row in organisasi_list.iterrows():
        new_ws = wb.copy_worksheet(ws)
        new_ws.title = row["short_name"]
        new_ws["A7"] = "Bulan: {} {}".format(get_nama_bulan(bulan), tahun)
        new_ws["A8"] = row["nama"]


if __name__ == "__main__":
    main("202401-001")
